from turtle import title
from numpy import void
import openpyxl
import pandas as pd
'''
excel处理器
'''
class ExcelHandler:
    def __init__(self):
        self.field=''
        self.context=[]
        pass

    def json2entity(self,json: dict) -> void:
        if 'field' in json and 'context' in json:
            self.field = json['field']
            self.context = json['context']
        else:
            raise ValueError("JSON data missing required fields")
    '''
    写入excel文件
    @param file_name: 文件名称
    @param data: json数据
    @param filterd: 筛选字段
    @param Pagination: 是否分页
    '''
    def write_excel(self,file_name,data,filterd=[],Pagination=False) -> void:
        self.json2entity(data)
        field = self.field
        data = self.context
        df_all = pd.DataFrame(data,columns=field)
        if(not str.endswith(file_name,".xlsx")):
            file_name = file_name + ".xlsx"
        
        '是否开启分页模式，默认精简模式'
        if Pagination and bool(filterd):
            '分页模式'
            df_filterd = df_all[filterd].copy()
            df_filterd.loc[:, '详情'] = df_filterd.index.map(lambda x: f'=HYPERLINK("#Sheet2!A{x+2}", "查看详情")')
            # 写入多个表单，使用 ExcelWriter
            with pd.ExcelWriter(file_name) as writer:
                df_filterd.to_excel(writer, sheet_name='Sheet1', index=False)
                df_all.to_excel(writer, sheet_name='Sheet2', index=False)

                ws = writer.book['Sheet1']
                from openpyxl.styles import Font
                # 找到“详情”列的列号
                for idx, cell in enumerate(ws[1]):
                    if cell.value == '详情':
                        detail_col = cell.column_letter
                        break
                # 从第二行开始遍历该列，设置样式
                for row in range(2, ws.max_row + 1):
                    cell = ws[f"{detail_col}{row}"]
                    cell.font = Font(color='0000FF', underline='single')
        else:
            '精简模式'
            # 将 DataFrame 写入 Excel 文件，写入 'Sheet1' 表单
            writer = pd.ExcelWriter(file_name, engine='openpyxl')
            df_all.to_excel(writer, sheet_name='Sheet1', index=False)
            worksheet = writer.sheets['Sheet1']
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter # Get the column name
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_name].width = adjusted_width
            writer.save()